import openpyxl
import os.path
import itertools
import datetime
import ls
import sync
import prnt

class D:
    def news_paper_name():
        return { 'ア':'朝日','マ':'毎日','ヨ':'読売','サ':'日本経済','産':'産経','ト':'東京','工':'日刊工業','日産':'日経産業','流通':'日経流通','報知':'スポーツ報知','デイリー':'デイリー  ','日刊スポ':'日刊スポーツ','A':'The Japan Times','Y':'The Japan News','F・T':'Financial Times','ヴェリタス':'ヴェリタス'} # 

class file():
    def active():
        file.rm('../paper/') # フォルダ内の全ファイル削除
        path = file.ph()
        create_wb = openpyxl.Workbook() # Excelファイルの新規作成
        create_wb.save( path)   # "内幸町8号朝刊.xlsx", ファイル生成
        tpl = prnt.LD_DL_LD.get() # {'い': [0, 1, 2, 3,], 'ろ': [23, 24, 25], 'は': [33]
        file.wr( path, tpl) # 書き込み

    def exhibition( wb,ws): # 展示
        ( bool(openpyxl.load_workbook("../paper/内幸町8号朝刊.xlsx")))
        # (wb["Sheet1"],wb.worksheets[0], wb.active)
        
        # ワークシートの全セルを反復（行ごと）
        for row in ws.iter_rows():
            for cell in row:
                (cell.value)
        
        # セルの値を取得する
        for row in ws.iter_rows(values_only=True):
            for value in row:
                (value)
                        
        # ワークシートの全セルを反復（列ごと）
        for column in ws.iter_cols(values_only=True):
            (column)
        
        for column in ws.columns:
            for cell in column:
                (cell.value)
        ws = wb.active  # アクティブなワークシートを選択
        (f'active sheet : {ws.title}')  # sheet name: Sheet
        # ワークシートの列挙
        for sheet in wb:
            (f'sheet : {sheet.title}')
        
    def ph():
        return "../paper/" + sync.dr.area() + ('朝刊' if sync.dr.d_d() != 'dusk' else '夕刊') + ".xlsx"
                    
    def rm( ph): # ファイル削除
        l = os.listdir( path = ph )
        [ os.remove( ph + '/' + i ) for i in l ] # ファイル削除
        
    def save( wb):
        path = file.ph() # ../paper/内幸町8号朝刊.xlsx
        wb.save(path) 
        # ファイルを閉じる
        wb.close()
    # 印刷設定
    def page( wb ): 
        print_area = 'A1:X600'
        for ws in wb.worksheets:
            # プリントエリアを設定
            ws.print_area = print_area
            # フッターを設定
            ws#.oddFooter.center.text = '&P / &Nページ'
            ws#.oddFooter.center.size = 15
        
            wps = ws.page_setup
            # 印刷の向きを設定
            wps.orientation = ws.ORIENTATION_PORTRAIT # （横：ORIENTATION_PORTRAIT）（縦：ORIENTATION_LANDSCAPE）
            # 横を１ページ
            wps.fitToWidth = 1
            # 縦を自動
            wps.fitToHeight = 0
            # fitTo属性を有効にする
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            # 用紙サイズを設定
            wps.paperSize = ws.PAPERSIZE_A4 # ws.PAPERSIZE_A5
            # A4サイズ「縦 210mm × 横 297mm」, B5サイズ「182mm × 257mm」
            wps.paperHeight = "297mm" # 単位：'mm' | 'cm' | 'in' | 'pt' | 'pc' | 'pi'
            wps.paperWidth  = "210mm" # 単位：'mm' | 'cm' | 'in' | 'pt' | 'pc' | 'pi'
            # 余白
            inch = 1 / 2.54 # 設定はインチ単位。 1インチ=2.54cm,  0.4インチで約1センチ,
            ws.page_margins.top    = inch
            ws.page_margins.right  = inch
            ws.page_margins.left   = inch
            ws.page_margins.footer = inch
            # 印刷品質 ※解像度
            ws.page_setup.horizontalDpi = 1200
            ws.page_setup.verticalDpi   = 1200
        
    def wr( path, tpl): # wirte :書き込み　
        path
        ld, dl, lst = tpl
        (ld,dl)
        wb = openpyxl.load_workbook( path) # Excelファイルの読み込み
        wb.remove( wb.worksheets[-1]) # 'sheet'を削除。
        (ld) # [{'区分': '人事院', '区間': '', '割当': 'い'}, {'区分': '弁護士会館', '区間': '', '割当': 'ろ'}, {'区分': '富国生命', '区間': 'A', '割当': 'は'}]
        (dl) # {'い': [0, 1, 2, 3,], 'ろ': [23, 24, 25], 'は': [33]
        alpha = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X']
        #
        # [{'区分': '富国生命', '区間': 'A', '割当': 'は'}, {'区分': '富国生命', '区間': 'C', '割当': 'は'}] {'は': [32, 34]}
        lld = LLD.LD_L_L( lst, dl, ls.kl(), [ 'メモ2','名称1','号数'])
        lll = LLL.LD_DL_L(  lst, dl, lld, alpha, [], []) # [[[9, 5, 9], [5, 4, 4, 5, 4], [9, 4, 10], [4, 5, 8, 4], [4, 9, 4], [9, 8, 4], [9, 6]], [[10, 7, 4],
        ('             ', lll)
        kivl = DL.LLL_L( lll, alpha, [], {}) # {0: [0, 24], 1: [0, 15], 2: [0, 10], 3: [1, 24],
        (kivl)
        dl1, dl2, arr, i2, i3, liii  = {}, {}, [], 2, 3, [ 50, 20, 25]
        li = cells.gyo_retsu( 4, i3) # メモ2, 縦書き,　行列, row colmun
        #
        for txt , l in dl.items():
            ws = wb.create_sheet( index= 0, title= txt)
            sheet.datetime( ws)
            sheet.meta( ws, txt, ld, alpha, liii, dl1, dl2 )
            sheet.footer(ws, ld, txt, dl1, dl2 )
            for i , n in enumerate(l) : # [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
                dic = lst[n]
                (file.exhibition( wb, ws))
                (i,n,dic.get('区分'),kivl.get(i))
                sheet.news( ws, lld, dic, alpha, liii, i2, li, i3, len(arr), kivl)
                arr.append(n)
                #
            sheet.formatting( wb, ws, l, alpha, i2) # 書式設定
        file.page( wb)
        file.save( wb)
            
class sheet():
    def footer( ws, ld, txt, dl1, dl2 ): 
        t1, t2, txt = S_S_S.LD_DL_DL_T( ld, dl1, dl2, txt) # 富国生命  , A , B , C は
        args = [ '区分:', t1,'     区間:', t2,'     割当:', txt,'     ページ', '&P / &N' ]
        # フッターを設定
        ws.oddFooter.center.text = ' '.join(args)  
        ws.oddFooter.center.size = 15
        
    def news( ws:any, lld, dic:dict, alpha:list, liii, i2:int, li:list, i3, n:int, kivl:dict):        
        ksvs = D.news_paper_name() # { 'ア':'朝日','マ':'毎日','ヨ':'読売','サ':'日本経済','産':'産経','ト':'東京','工':'日刊工業','日産':'日経産業','流通':'日経流通','報知':'スポーツ報知','A':'The Japan Times','Y':'The Japan News','F・T':'Financial Times'}
        i2 # 2 , 起点, 列, 顧客情報
        
        dct = { k : v for k , v in dic.items() if k not in ls.kl() }  # {'ア': '4', 'マ': '5', 'ヨ': '3', 'サ': '3', '産': '4', 'ト': '3'},  {'ア': '1', 'サ': '2'}
        d1 = { k : dct.get(k) for k in list(dct)[:21]} # 新聞名, 最大表示数
        d2 = { i : dic.get(i) for i in ['メモ2','名称1','号数'] } # { '名称1': '秘書室①', '号数': '1', 'メモ2': ''}
        dct = { **d1, **d2 } # {'ア': '2', 'マ': '2', 'ヨ': '2', 'サ': '3', '産': '2', 'ト': '2', 'メモ2': '', '名称1': '秘書室①', '号数': '1'}
        #
        r_ , c_ = kivl.get( n ) # {0: [0, 24], 1: [0, 15], 2: [0, 10], 3: [1, 24], 4: [1, 19], 5: [1, 15], 6: [1, 11], 7: [1, 6], 8: [2, 24], 9: [2, 15], 10: [2, 11], 11: [3, 24], 12: [3, 20], 13: [3, 15], 14: [3, 7], 15: [4, 24], 16: [4, 20], 17: [4, 11], 18: [5, 24], 19: [5, 15], 20: [5, 7], 21: [6, 24], 22: [6, 15], 23: [0, 24], 24: [0, 14], 25: [0, 7], 26: [1, 24], 27: [1, 11], 28: [1, 4], 29: [2, 24], 30: [2, 16], 31: [2, 11], 32: [0, 24]}
        ( n , r_ , c_ , len(dct))
        #
        for i , tpl in enumerate( dct.items()):
            k , v = tpl
            r = 3 * r_ + i2 
            c = c_ - i
            ( r , c )
            line = k if k not in ksvs else ksvs.get( k )
            
            if k not in ['メモ2','名称1','号数'] :
                # 1行目
                cells.ppty( ws,  r , c , line , { 'slice':15, 'size':sheet.fonts( line).get('size'), 'name':sheet.fonts( line).get('name'), 'wrapText': sheet.algm( k, line ).get('wrapText'), 'shrink_to_fit':True, 'textRotation':sheet.algm( k, line ).get('textRotation'), 'horizontal':sheet.algm( k, line ).get('horizontal'), 'vertical':sheet.algm( k, line ).get('vertical'), 'top':'thick', 'bottom':'thick', 'left':'thick', 'right':'thick', 'end_row': r , 'end_column': c })
                # 2行目
                cells.ppty( ws,  r + 1 , c , v , { 'slice':2, 'size':18, 'bold':True, 'shrink_to_fit':True, 'horizontal':'center', 'vertical':'center', 'top':'thin', 'bottom':'thick', 'left':'thick', 'right':'thick', 'end_row':r + 1, 'end_column': c })
                
            elif k in ['メモ2'] :
                cells.ppty( ws, r , sheet.c_( c, alpha, dct, [ '名称1', '号数'] ) -1 , cells.tategaki( v, li, i3 ), { 'size':15, 'wrapText':True,  'horizontal':'center', 'vertical':'center', 'end_row':r + 1, 'end_column': sheet.c_( c, alpha, dct, [ '名称1','号数'] ) })
            
            elif k in ['名称1'] :
                # 3行目
                cells.ppty( ws, r + 2 , c + len(d1) , v , { 'slice':7, 'size':14, 'shrink_to_fit':True, 'horizontal':'center', 'vertical':'center', 'end_row':r + 2, 'end_column':c + len(d1) + 1})

            elif k in ['号数'] :
                # 3行目
                cells.ppty( ws, r + 2 , c + len(d1) , v , { 'slice':10, 'size':14, 'bold':True, 'shrink_to_fit':True, 'horizontal':'center', 'vertical':'center', 'top':'thin', 'bottom':'thick', 'left':'thick', 'right':'thick', 'end_row':r + 2, 'end_column':c + len(d1)})
        # ボーダー
        for i in range( 1,len(alpha)) :
            ws.cell( row = r + 2, column =  i ).border = openpyxl.styles.borders.Border( bottom = openpyxl.styles.borders.Side(style='dotted'), ) 
        ws.row_dimensions[ r + 2 ].border #= openpyxl.styles.borders.Border( bottom = openpyxl.styles.borders.Side(style='thin'), )
        # row, 寸法
        sheet.dimension( ws , r, liii ) 
                
    def dimension( ws , r , liii): # 寸法 ,  縦寸法：縦 "297mm", 1140(95 * 12) + 余白:2インチ
        ws.row_dimensions[ r ].height = liii[0] # 1行目, 50 = 1.75cm 100= 3.52cm 10 = 0.34cm
        ws.row_dimensions[ r + 1 ].height = liii[1] # 2行目 20 = 0.69cm, 10 = 0.345cm
        ws.row_dimensions[ r + 2 ].height = liii[2]# 3行目
        
    def formatting( wb:any, ws:any, l:list, alpha:list, ix:int): # 書式設定: formatting
        # 列の幅を調整(幅)
        for i in alpha:
            ws.column_dimensions[i].width = 6
        # 行の高さを調整
        for i in range( ix, 30): # 行の範囲
            if i % 2 == 0 and i >= ix : # 偶数, 行, 新聞名
                ws.row_dimensions[i].height# = 50
            else:
                ws.row_dimensions[i].height# = 20
        lt = [ ( c , r ) for r , c in itertools.product( range( 1, len( alpha)+1 ), range( 2, 40 ) )] # c, r
        # フォントサイズを指定
        for r , c in lt:
            ws.cell( row=r, column=c )#.font = openpyxl.styles.Font( size= 14) 
        # 
        for r , c in lt : # 999999, 808080, ffff00, c0c0c0, 0000ff, 008000
            ws.cell( row= r , column= c )#.border = openpyxl.styles.borders.Border(top=openpyxl.styles.borders.Side(style='thick'), bottom=openpyxl.styles.borders.Side(style='thick'), left=openpyxl.styles.borders.Side(style='thick'), right=openpyxl.styles.borders.Side(style='thick'))

    def datetime( ws ):
        ws['U1'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M') 
        ws['U1'].font = openpyxl.styles.Font( size= 16 ) 
        ws['U1'].alignment = openpyxl.styles.Alignment( horizontal = 'center', vertical = 'center') 
        ws.merge_cells( start_row= 1 , end_row= 1, start_column= 21, end_column=  24 )

    def meta( ws, txt, ld, alpha, liii, dl1, dl2):
        txt # 'い'
        num = 52        
        t1, t2, txt = S_S_S.LD_DL_DL_T( ld, dl1, dl2, txt)
        sheet.obj( ws, { 'k':'区分', 'v':t1,  'r': 1, 'c': 1,   'start_column': 1 + 1 ,    'end_column': 1 + 10, 'fontsize': num })
        sheet.obj( ws, { 'k':'区間', 'v':t2,  'r': 1, 'c': 12,  'start_column': 12 + 1,    'end_column': 12 + 4, 'fontsize': num })
        sheet.obj( ws, { 'k':'割当', 'v':txt, 'r': 1, 'c': 17,  'start_column': 17 + 1,    'end_column': 17 + 2, 'fontsize': num - 14 })
        # 
        for i in range( 1, len(alpha)) :
            ws.cell( row = 1, column =  i ).border      = openpyxl.styles.borders.Border( bottom = openpyxl.styles.borders.Side(style='dotted'), ) 
            ws.cell( row = 1, column =  i ).alignment   = openpyxl.styles.Alignment( shrink_to_fit = True, horizontal = 'left', vertical = 'center') # shrink_to_fit:文字の自動サイズ調整
        ws.row_dimensions[1].height = sum(liii) # num             

    def obj( ws , dct): # セル
        k , v , r , c , start, end, fontsize = dct.values()
        ws.cell( row = r , column = c ).value = dct.get('k')
        ws.cell( row = r , column = c + 1 ).value = dct.get('v')
        ws.cell( row = r , column = c + 1 ).font = openpyxl.styles.Font( size = dct.get('fontsize')  ) 
        ws.cell( row = r , column = c + 1 ).alignment = openpyxl.styles.Alignment( shrink_to_fit = True, horizontal = 'left', vertical = 'center') # shrink_to_fit : 文字の自動サイズ調整         # 縦書き:textRotation = 255, 横位置：horizontal ,縦位置：vertical ,縦書きにshrink_to_fit・wrapText効果なし、
        ws.merge_cells( start_row = r , end_row = r , start_column = dct.get('start_column')  , end_column = dct.get('end_column') )
        
    def c_( c, alpha, dct, args)->int:
        alpha # ['A','B','C','D','E','F','G',
        args # ['名称1', '号数']
        len(alpha) - len(args) # 22
        if len(dct) == 1:
            c = c - 2
        elif len(dct)==2:
            c = c - 1
        else :
            c = c
        return c 
        
    def fonts( line): # size_font
        if len(line) == 2:
            dct = {'size':18,'name':'BIZ UDP明朝 Medium'}
        elif len(line) == 3:
            dct = {'size':14,'name':'BIZ UDP明朝 Medium'}
        elif len(line) == 4:
            dct = {'size':12,'name':'BIZ UDゴシック'}
        elif len(line) == 5:
            dct = {'size':12,'name':'BIZ UDP明朝 Medium'}
        elif len(line) > 5:
            dct = {'size':11,'name':'Cambria'}
        else :
            dct = {'size':15,'name':'BIZ UDP明朝 Medium'}
        return dct
    
    def algm( txt, line ): # Alignment: アラインメント:一直線にすること,整列
            if txt  in ls.nw()[0] :
                # 縦書き:textRotation = 255, 横位置：horizontal ,縦位置：vertical ,縦書きにshrink_to_fit・wrapText効果なし、
                dct = { 'wrapText': None, 'textRotation': 255, 'horizontal': 'center', 'vertical': 'bottom'}
            elif len(line) <= 4:
                dct = { 'wrapText': None, 'textRotation': 255, 'horizontal': 'center', 'vertical': 'bottom'}
            else:
                dct = { 'wrapText': True, 'textRotation': 90, 'horizontal':  'center', 'vertical': 'center'}
            return dct
        
class cells:
    def tategaki( txt, li, i3:int ):
        ( txt, li, len(li))
        # [ 8, 4, 0, 9, 5, 1, 10, 6, 2, 11, 7, 3 ]
        line = txt[ :len( li)]
        for i in range( len(list(enumerate(line))), len(li)): # 
            line += '　' # 全角空白を追加, liと文字数を合わせる,
        ( line, len(line), )
        ls = [ '' for i in range( 0, len(line))]
        for i , x in enumerate(line):
            ( i , x)
            ls[i] = line[li[i]]
            if i in [ i -1 for i in range( 1, len(li)) if i % i3 in [0]]: # [2, 5, 8, 11]
                ls[i] = ls[i] + '\n'
        (ls)
        string = ''.join(ls)
        return string
    
    # メモ2, 縦書き,　行列, row colmun
    def gyo_retsu( gyo:int, retsu:int)->list: 
        # 840
        # 951
        # A62
        # B73
        [ 4*2+0, 4*1+0, 4*0+0, 4*2+1, 4*1+1, 4*0+1, 4*2+2, 4*1+2, 4*0+2, 4*2+3, 4*1+3, 4*0+3 ] 
        row , column = gyo, retsu # 4 , 3
        l = [ row * I + i for i in range( 0 , row ) for I in range( 0 , column)[::-1] ] 
        return l # [ 8, 4, 0, 9, 5, 1, 10, 6, 2, 11, 7, 3 ]
    
    def ppty( ws, r , c , v , d , ): # property:属性
        (d.get('wrapText'),type(d.get('wrapText')))
        ws.cell( row = r , column = c ).value = v[:d.get('slice')]
        ws.cell( row = r , column = c ).font = openpyxl.styles.Font( size= d.get('size'), name= d.get('name'), bold= d.get('bold')) 
        ws.cell( row = r , column = c ).alignment = openpyxl.styles.Alignment( shrink_to_fit= d.get('shrink_to_fit'), textRotation = d.get('textRotation'), horizontal = d.get('horizontal'), vertical = d.get('vertical'), wrapText = d.get('wrapText') ) # 
        ws.cell( row = r , column = c ).border = openpyxl.styles.borders.Border( top= openpyxl.styles.borders.Side(style=d.get('top')), bottom= openpyxl.styles.borders.Side(style=d.get('bottom')), left= openpyxl.styles.borders.Side(style=d.get('left')), right= openpyxl.styles.borders.Side(style=d.get('right')))
        ws.merge_cells( start_row = r , end_row = d.get('end_row') , start_column = c , end_column = d.get('end_column') )
        
class LLD:
    def LD_L_L( lst:list, dl, keylist:list, lis:list):
        keylist # ls.kl()
        lis # ['メモ2','名称1','号数']
        lld = [ [ { **{ k : v for k , v in lst[i].items() if k not in keylist }
        , **{ k : v for k , v in lst[i].items() if k in lis[0] }
        , **{ k : v for k , v in lst[i].items() if k in lis[1] }
        , **{ k : v for k , v in lst[i].items() if k in lis[2] }} for i in l ] for txt, l in dl.items() ] #[[{'名称1': '秘書室①', '号数': '1', 'メモ2': '', 'ア': '2', 'マ': '2', 'ヨ': '2', 'サ': '2', '産': '2', 'ト': '2'},
        return lld # [[{'ア': '2', 'マ': '2', 'ヨ': '2', 'サ': '3', '産': '2', 'ト': '2', 'メモ2': '', '名称1': '秘書室①', '号数': '1'}, {'ア': '1', 'サ': '2', 'メモ2': '', '名称1': '総務課②', '号数': '２'},

class I:
    def LLD(lld): # 要素の総和
        l = [ len([ [ i for i in  I ] for I in ll ]) for ll in lld ]
        return sum( l )
        [ l for ll in lld for l in ll ]
        
    def LLL(lll):  # 要素の総和
        l = [ i for ll in lll for l in ll for i in l ]
        return len(l)
    
class S_S_S:
    def LD_DL_DL_T( ld:list, dl1:dict, dl2:dict, txt:str)->str :
        dl1.setdefault( txt, [])
        dl2.setdefault( txt, [])
        for d in ld: # [{'区分': '人事院', '区間': '', '割当': 'い'}, {'区分': '弁護士会館', '区間': '', '割当': 'ろ'},
            if txt == d.get('割当'): # 'い' == 'い'
                dl1[txt].append( d.get('区分'))
                dl2[txt].append( d.get('区間'))
        t1 =  ",".join(  set(dl1[txt]))
        t2 =  " , ".join(sorted(set(dl2[txt])))
        return t1, t2, txt #  # 富国生命  , A , B , C は
    
class LL:
    def LLD(lld):
        ll = [ [ len({ k : v for k , v in d.items() }) for d in ll ] for  ll in lld ]
        return ll # [[9, 5, 9, 5, 4, 4, 5, 4, 9, 4, 10, 4, 5, 8, 4, 4, 9, 4, 9, 8, 4, 9, 6], [10, 7, 4, 13, 7, 4, 8, 5, 9], [8]]

    def S_I( string, n, ll): # ABCDEFGHIJKLMNOPQRSTUVWXYZ, 5 ,[]
        ls = []
        last_index = list( range(len(string)))[ -1] # 25
        for i , x in enumerate(string):
            if len( ls) <  n:
                ls.append( x )
                if len(ls) >=  n:
                    ll.append( ls )
                    ls = []
                elif i == last_index :
                    ll.append( ls )
                    ls = []
        return ll # [['A', 'B', 'C', 'D', 'E'], ['F', 'G', 'H', 'I', 'J'], ['K', 'L', 'M', 'N', 'O'], ['P', 'Q', 'R', 'S', 'T'], ['U', 'V', 'W', 'X', 'Y'], ['Z']]
    
class LLL:
    # alphaの要素数ごとに区切る
    def LD_DL_L( lst:list, dl:dict, lld, alpha:list, arr:list,  lll:list):
        ll = LL.LLD(lld) # [[9, 5, 9, 5, 4, 4, 5, 4, 9, 4, 10, 4, 5, 8, 4, 4, 9, 4, 9, 8, 4, 9, 6], [10, 7, 4, 13, 7, 4, 8, 5, 9], [8]]
        ('         ', ll )
        ll[0] # [9, 5, 9, 5, 4, 4, 5, 4, 9, 4, 10, 4, 5, 8, 4, 4, 9, 4, 9, 8, 4, 9, 6]
        for l in ll :
            arry = []
            for i , n in enumerate( l ) :
                (i) # 0 1 2 3 4 5 6 7 8 9 10 11
                (n) # 9 5 9 5 4 4 5 4 9 4 10 4 5 8 4 4 9 4 9 8 4 9 6
                if sum( arr ) + n <= len(alpha):
                    arr.append( n )
                    last_index = list( range(len(l)))[ -1] # 最期、 0
                    ( i , sum(arr) , last_index )
                    if i == last_index :
                        arry.append( arr ) # [[9, 5, 9], [5, 4, 4, 5, 4], [9, 4, 10], [4, 5, 8, 4], [4, 9, 4], [9, 8, 4], [9, 6]]
                        (arry)
                        arr = []
                elif sum( arr ) + n > len(alpha):
                    arry.append( arr ) # [[9, 5, 9], [5, 4, 4, 5, 4]]
                    arr = []
                    arr.append(n)
                    if i == last_index :
                        arry.append( arr ) # [[9, 5, 9], [5, 4, 4, 5, 4], [9, 4, 10], [4, 5, 8, 4], [4, 9, 4], [9, 8, 4], [9, 6]]
                        (arry)
                        arr = []
            lll.append(arry)    
        ( I.LLD(lld) == I.LLL(lll) ) # 238 ==238
        return lll # [[[9, 5, 9], [5, 4, 4, 5, 4], [9, 4, 10], [4, 5, 8, 4], [4, 9, 4], [9, 8, 4], [9, 6]],
       
class DL:
    # column値を算出, 最大値24からいくつの値か,
    def LLL_L( lll:list, alpha:list, arr:list, kivl:dict)-> dict :
        for ll in lll : # [[[9, 5, 9], [5, 4, 4, 5, 4], [9, 4, 10], [4, 5, 8, 4], [4, 9, 4], [9, 8, 4], [9, 6]],
            for I, l in enumerate(ll) : # [[9, 5, 9], [5, 4, 4, 5, 4]]
                for i , n in enumerate(l): # 0, [9, 5, 9]
                    ( len(arr), i, sum(l), len(alpha) - sum( l[:i]), n, l, I) # 
                    (l[:i]) # []
                    kivl.setdefault( len(arr) , [])
                    kivl[ len(arr) ] = [ I , len(alpha) - sum( l[:i])] # 24 - sum( [9, 5, 9][:0])
                    arr.append(i)
        (kivl)
        return kivl # {0: [0, 24], 1: [0, 15], 2: [0, 10], 3: [1, 24],
    
class L: 
    # Iの範囲で、 nの倍数をリストで返す, 
    def I_I( I:int, n:int)->list:
        l = [ i for i in range( 1, I) if i % n in [0]]
        ( l )
        return l
       
if __name__ == '__main__': # これがないと外部からインポートされた際に処理が実行されてしまう。    
    file.active()
    lst = ls.customer()
    dic = lst[0]
    t0, t1, t2, t3, t4, t5, t6 = [ dic.get(i) for i in ls.tm()]
    ( t0, t1, t2, t3, t4, t5, t6)
    
    tpl = prnt.LD_DL_LD.get() # {'い': [0, 1, 2, 3,], 'ろ': [23, 24, 25], 'は': [33]
    ld, dl, l = tpl
    lld = [ [{ k : v for k , v in lst[i].items() if k not in ls.kl()} for i in l] for txt, l in dl.items()] # [[{'ア': '2', 'マ': '2', 'ヨ': '2', 'サ': '3', '産': '2', 'ト': '2'}, {'ア': '1', 'サ': '2'}, {'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '1', '産': '1', 'ト': '1'}, {'ア': '1', 'ヨ': '1'}, {'サ': '1'}, {'サ': '1'}, {'サ': '1', 'A': '1'}, {'サ': '1'}, {'ア': '2', 'マ': '1', 'ヨ': '3', 'サ': '3', '産': '1', 'ト': '1'}, {'サ': '1'}, {'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '1', '産': '1', 'A': '1', '農業': '1'}, {'サ': '1'}, {'ヨ': '1', 'サ': '1'}, {'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '3', 'ト': '1'}, {'サ': '1'}, {'ヨ': '1'}, {'ア': '1', 'マ': '1', 'ヨ': '2', 'サ': '2', '産': '1', 'ト': '1'}, {'サ': '1'}, {'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '3', '産': '1', 'ト': '1'}, {'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '1', '産': '1'}, {'産': '1'}, {'ア': '2', 'マ': '2', 'ヨ': '2', 'サ': '2', '産': '2', 'ト': '2'}, {'マ': '1', 'ヨ': '1', 'サ': '1'}], [{'ア': '4', 'マ': '4', 'ヨ': '4', 'サ': '5', '産': '3', 'ト': '4', 'A': '1'}, {'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '1'}, {'サ': '1'}, {'ア': '2', 'マ': '2', 'ヨ': '2', 'サ': '2', '産': '2', 'ト': '2', 'A': '1', '報知': '1', 'スポニチ': '1', '碁': '1'}, {'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '1'}, {'サ': '1'}, {'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '1', '日刊スポ': '1'}, {'ヨ': '1', 'サ': '1'}, {'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '1', '産': '1', 'ト': '1'}], [{'ア': '1', 'マ': '1', 'ヨ': '1', 'サ': '1', '産': '1'}]]
    (lld)
    lll = [[[9, 5, 9], [5, 4, 4, 5, 4], [9, 4, 10], [4, 5, 8, 4], [4, 9, 4], [9, 8, 4], [9, 6]], [[10, 7, 4], [13, 7, 4], [8, 5, 9]], [[8, 5], [13, 5], [9, 4, 7], [9, 4, 4, 5], [5, 9, 4, 4], [4, 5, 4, 4, 5], [4, 4, 5, 4, 4], [6, 4, 4, 5, 4], [4, 4, 4, 4, 4, 4], [4, 10, 4, 5], [4, 6, 4]], [[22], [8, 13], [8, 8, 8], [6, 8, 9], [4, 4, 5, 4, 4], [4, 5, 11, 4], [5, 4, 5, 4], [8, 4, 11], [6, 8, 7], [4, 11, 7]], [[9, 9], [9, 9], [9, 10], [6, 9, 9], [7, 9, 7], [11, 9], [11, 11], [12, 12], [13, 9], [8, 11], [9, 10], [11, 9], [10, 9], [9, 9], [9, 11], [8, 7], [11, 9], [8, 11], [10, 7], [9, 9], [9, 9], [9, 9], [9, 9], [9, 8, 4], [6, 11], [9, 9, 6], [9, 4]], [[4, 5, 6, 4, 4], [4, 5, 5, 5, 4], [11, 4], [10, 4, 6], [11, 5, 4, 4], [4, 4, 6, 9], [4, 4, 9, 4], [4, 6, 5, 4, 4], [5, 4, 4, 5], [10, 4, 4, 4], [4, 4, 5, 4, 4], [4, 9, 4, 4], [4, 6, 4, 5, 5], [5, 4, 4, 4, 6], [4]], [[4, 4, 5, 7, 4], [12, 4, 4], [5, 9, 4, 4], [4, 4]]]
    [ [ [ i for i in  I ] for I in ll ]  for ll in lll ]
    alpha = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    LL.S_I( ''.join(alpha), 5, [])
    dict(enumerate( 'hello world!')).keys() # dict_keys([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]) 
    cells.tategaki( 'hello world!', [ 8, 4, 0, 9, 5, 1, 10, 6, 2, 11, 7, 3 ], 3)  #
    L.I_I( 20 , 3) # nの倍数をリストで返す, Iの範囲で、
    li = map(lambda x: x -1 , [3, 6, 9, 12] ) # [2, 5, 8, 11], 全ての要素に処理をする
    (list(li))