import openpyxl # sheet
import os.path
import itertools
import datetime
import ls
import sync
import prnt
import mgmt

class file():
    def active():
        file.rm('../table/') # フォルダ内の全ファイル削除
        path, area, d_d, lst = file.ph(), sync.dr.area(), sync.dr.d_d(), ls.customer()
        create_wb = openpyxl.Workbook() # Excelファイルの新規作成
        create_wb.save( path)   # "内幸町8号朝刊.xlsx", ファイル生成
        file.strc( path, area, d_d, lst) # 書き込み

    def ph():
        return '../table/' + sync.dr.area() + ('朝刊' if sync.dr.d_d() == 'dawn' else '夕刊') + ".xlsx"
                    
    def rm( ph): # ファイル削除
        l = os.listdir( path = ph )
        [ os.remove( ph + '/' + i ) for i in l ] # ファイル削除
        
    def save( wb):
        path = file.ph() # /内幸町8号朝刊.xlsx
        wb.save(path) 
        # ファイルを閉じる
        wb.close()
    # 印刷設定
    def page( wb, letter, n): 
        print_area = 'A1:AG99'
        print_area = 'A1:'+ letter + str(n) # AG 9'
        for ws in wb.worksheets:
            # プリントエリアを設定
            ws.print_area = print_area
        
            wps = ws.page_setup
            # 印刷の向きを設定
            wps.orientation = ws.ORIENTATION_PORTRAIT # 横：ORIENTATION_PORTRAIT 縦：ORIENTATION_LANDSCAPE
            # 横を１ページ
            wps.fitToWidth = True
            # 縦を自動
            wps.fitToHeight = True
            # fitTo属性を有効にする
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            # 用紙サイズを設定
            wps.paperSize = ws.PAPERSIZE_A4 # ws.PAPERSIZE_A5
            # A4サイズ「縦 210mm × 横 297mm」, B5サイズ「182mm × 257mm」
            wps.paperHeight = "210mm" # 単位：'mm' | 'cm' | 'in' | 'pt' | 'pc' | 'pi'
            wps.paperWidth  = "297mm" # 単位：'mm' | 'cm' | 'in' | 'pt' | 'pc' | 'pi'
            # 余白
            inch = 1 / 2.54 # 設定はインチ単位。 1インチ=2.54cm,  0.4インチで約1センチ,
            ws.page_margins.top    = inch * 2
            ws.page_margins.right  = inch
            ws.page_margins.left   = inch
            ws.page_margins.bottom = inch * 2
            ws.page_margins.header = inch
            ws.page_margins.footer = inch
            # 印刷品質 ※解像度
            ws.page_setup.horizontalDpi = 1200
            ws.page_setup.verticalDpi   = 1200
        
    def strc( path:str, area:str, d_d:str, lst:list): # structure:構造
        path, area # ../table/内幸町8号朝刊.xlsx, 内幸町8号
        hss = mgmt.DD.LD( lst, {}) # {'人事院': {'ア': 14, 'マ': 12, 'ヨ': 18, 'サ': 31, '産': 11, 'ト': 9, 'A': 2, '農業': 1}, '弁護士会館': {'ア': 10, 'マ': 10, 
        ktvd = mgmt.Dtd.LD_L(lst, ['区分', '区間'], {}) # {('人事院', ''): {'ア': 14, 'マ': 12, 'ヨ': 18, 'サ': 31, '産': 11, 'ト': 9, 'A': 2, '農業': 1}, ('弁護士会館', ''): {'ア': 10,
        sst = { k for d in lst for k in d.keys() if k not in ls.kl() } # {'Y', '日刊スポ', 'ガイド', '産', '海事', 'A', 
        dct = { k : None for k in mgmt.L.news() if k in sst } # {'ト': None, '産': None, 'サ': None, 'ヨ': None, 'マ': None, 'ア': None, '工': None, '日産': None,
        d_d = '朝刊' if sync.dr.d_d() == 'dawn' else '夕刊'
        aso = D.DD( ktvd, {})# Associative array: 連想配列
        ( path, area )
        wb = openpyxl.load_workbook( path) # Excelファイルの読み込み
        wb.remove( wb.worksheets[-1]) # 'sheet'を削除。        #
        ws = wb.create_sheet( index= 0, title= area + d_d)
        sheet.header( ws, [ area, d_d ])
        
        # セル番地を取得, セル行を取得,  A1 1 A 
        ( ws.cell(row=1,column=1).coordinate, ws.cell( row=1, column=1).column, openpyxl.utils.get_column_letter(1))
        x = ws.cell( row= 1, column= 1).column 
        sheet.news( ws, ktvd, dct, aso, x)
        sheet.total( ws, ktvd, dct, aso, x)
        
        # セル番地を取得 'A1'のボーダーを設定。
        ws.cell( row = x , column = x ).border = openpyxl.styles.borders.Border( top= openpyxl.styles.borders.Side(style='medium'), bottom= openpyxl.styles.borders.Side(style='medium'), left= openpyxl.styles.borders.Side(style='medium'), right= openpyxl.styles.borders.Side(style='medium'))
        # 行の高さ, 1行目
        ws.row_dimensions[ x ].height = 80
        # 列幅, 1列目
        ws.column_dimensions['A'].width = 17
        # 最後の列の記号、　行の個数、　
        letter, n = [ openpyxl.utils.get_column_letter( i ) for i in range( 1, len(dct) + x + 1 )][-1], len(ktvd) + x + 1
        file.page( wb, letter, n) # 印刷設定
        file.save( wb)
            
class sheet():    
    def news( ws:any, ktvd, dct:dict, aso, x ):        
        ksvs = D.news_paper_name()
        (dct ) # {'ト': None, '産': None, 'サ': None, 'ヨ': None, 'マ': None, 'ア': None, '工': None,
        aso # {'ア': 270, 'マ': 247, 'ヨ': 268, 'サ': 358,
        for tpl , ksvi in ktvd.items():
            kb, kk = tpl
            kbkk = kb + ' ' + kk # 人事院 A
            ( kbkk, ksvi) # 人事院 {'ア': 14, 'マ': 12, 'ヨ': 18, 'サ': 37, '産': 11, 'ト': 9, 'A': 2, '農業': 1}
            r = list( ktvd.keys()).index(tpl) + x + 1
            # 行、 項目
            cells.ppty( ws, r , 1 , kbkk , { 'slice':10, 'size':14, 'shrink_to_fit':True, 'horizontal':'left', 'vertical':'center', 'top':'medium', 'bottom':'medium', 'left':'medium', 'right':'medium', 'end_row': r , 'end_column': 1 })
            # 総計
            for i , txt in enumerate(dct): 
                c = i + 2
                # 列、項目
                cells.ppty( ws, x , c , txt.replace( 'ー', '丨') if 'ー' in txt else txt, { 'slice': 5, 'size': 14, 'horizontal': 'center', 'vertical': 'bottom', 'top': 'medium', 'bottom': 'medium', 'left': 'medium', 'right': 'medium', 'end_row': x, 'end_column': c , 'wrapText': None, 'textRotation': 255 })
                # 値
                v = ktvd[ tpl].get(txt)
                v = v if bool(v) else ''
                cells.ppty( ws, r , c , v , { 'slice':3, 'size':18, 'shrink_to_fit':True, 'horizontal':'center', 'vertical':'center', 'top':'thin', 'bottom':'thin', 'left':'thin', 'right':'thin', 'end_row':r , 'end_column': c })

    # 総計
    def total(ws, ktvd, dct, aso, x):
        z = len(ktvd) + x + 1 # shrink_to_fit:「縮小して全体表示」
        cells.ppty( ws, z , x , '総計' , { 'slice':10, 'size':14, 'shrink_to_fit':True, 'horizontal':'center', 'vertical':'center', 'top':'medium', 'bottom':'thin', 'left':'thin', 'right':'thin', 'end_row': z , 'end_column': 1 })
        for i , txt in enumerate(dct): 
            c = i + x + 1
            cells.ppty( ws, z , c , aso.get(txt) , { 'slice':3, 'size':18, 'shrink_to_fit':True, 'horizontal':'center', 'vertical':'center', 'top':'thin', 'bottom':'thin', 'left':'thin', 'right':'thin', 'end_row': z , 'end_column': c })
            sheet.column_width( ws, c, aso.get(txt))
            
    # セルの値の桁数によって列幅を変更する, 100:3桁:width+3          
    def column_width( ws:any, c:int, n :int):
        s = openpyxl.utils.get_column_letter( c ) # セル記号を取得す, A
        (n,len(str(n)))
        ws.column_dimensions[ s ].width = 3 + len(str( n ))   
         
    # ヘッダーを設定
    def header( ws, args ): 
        l = ['【部数一覧】'] + args 
        ws.oddHeader.left.text = ' '.join(l)  
        ws.oddHeader.left.size = 16
        ws.oddHeader.right.text = datetime.datetime.now().strftime('%Y-%m-%d %H:%M') 
        ws.oddHeader.right.size = 16
                
class cells:        
    def ppty( ws, r , c , v , d , ): # property:属性
        ( d.get('wrapText'), type(d.get('wrapText')))
        ws.cell( row = r , column = c ).value = str(v)[:d.get('slice')]
        ws.cell( row = r , column = c ).font = openpyxl.styles.Font( size= d.get('size'), name= d.get('name'), bold= d.get('bold')) 
        ws.cell( row = r , column = c ).alignment = openpyxl.styles.Alignment( shrink_to_fit= d.get('shrink_to_fit'), textRotation = d.get('textRotation'), horizontal = d.get('horizontal'), vertical = d.get('vertical'), wrapText = d.get('wrapText') ) # 
        ws.cell( row = r , column = c ).border = openpyxl.styles.borders.Border( top= openpyxl.styles.borders.Side(style=d.get('top')), bottom= openpyxl.styles.borders.Side(style=d.get('bottom')), left= openpyxl.styles.borders.Side(style=d.get('left')), right= openpyxl.styles.borders.Side(style=d.get('right')))
        ws.merge_cells( start_row = r , end_row = d.get('end_row') , start_column = c , end_column = d.get('end_column') )
        ws.row_dimensions[ r ].height = 20 # 行の高さ

    
class D:
    def news_paper_name():
        return { 'ア':'朝日','マ':'毎日','ヨ':'読売','サ':'日本経済','産':'産経','ト':'東京','工':'日刊工業','日産':'日経産業','流通':'日経流通','報知':'スポーツ報知','デイリー':'デイリー  ','日刊スポ':'日刊スポーツ','A':'The Japan Times','Y':'The Japan News','F・T':'Financial Times','ヴェリタス':'ヴェリタス'} # 

    def DD( ktvd:dict, dct): # {('人事院', ''): {'ア': 14, 'マ': 12, 'ヨ': 18, 'サ': 31, '産': 11, 'ト': 9, 'A': 2, '農業': 1}, ('弁護士会館', ''): 
        for d in ktvd.values():
            for k , v  in d.items(): 
                dct.setdefault( k , 0)
                dct[ k ] += int( v )
        return dct # {'ア': 270, 'マ': 247, 'ヨ': 268, 'サ': 358,
    
if __name__ == '__main__': # これがないと外部からインポートされた際に処理が実行されてしまう。    
    file.active()
    lst = ls.customer()
    dic = lst[0]
    t0, t1, t2, t3, t4, t5, t6 = [ dic.get(i) for i in ls.tm()]
    ( t0, t1, t2, t3, t4, t5, t6)
    
    tpl = prnt.LD_DL_LD.get() # {'い': [0, 1, 2, 3,], 'ろ': [23, 24, 25], 'は': [33]
    ld, dl, l = tpl
    lld = [ [{ k : v for k , v in lst[i].items() if k not in ls.kl()} for i in l] for txt, l in dl.items()] # [[{'ア': '2', 'マ': '2', 'ヨ': '2', 'サ': '3', '産': '2', 'ト': '2'}, {'ア': '1', 'サ': '2'},
    (lld)
    lll = [[[9, 5, 9], [5, 4, 4, 5, 4], [9, 4, 10], [4, 5, 8, 4], [4, 9, 4], [9, 8, 4], [9, 6]], [[10, 7, 4], [13, 7, 4],
            [8, 5, 9]], [[8, 5], [13, 5], [9, 4, 7], [9, 4, 4, 5], [5, 9, 4, 4], [4, 5, 4, 4, 5], [4, 4, 5, 4, 4], [6, 4, 4, 5, 4],
            [4, 4, 4, 4, 4, 4], [4, 10, 4, 5], [4, 6, 4]], [[22], [8, 13], [8, 8, 8], [6, 8, 9], [4, 4, 5, 4, 4], [4, 5, 11, 4], [5, 4, 5, 4], [8, 4, 11], [6, 8, 7], [4, 11, 7]], 
            [[9, 9], [9, 9], [9, 10], [6, 9, 9], [7, 9, 7], [11, 9], [11, 11], [12, 12], [13, 9], [8, 11], [9, 10], [11, 9], [10, 9], [9, 9], [9, 11], [8, 7], [11, 9], [8, 11], [10, 7],
             [9, 9], [9, 9], [9, 9], [9, 9], [9, 8, 4], [6, 11], [9, 9, 6], [9, 4]], [[4, 5, 6, 4, 4], [4, 5, 5, 5, 4], [11, 4], [10, 4, 6], [11, 5, 4, 4], [4, 4, 6, 9], [4, 4, 9, 4], [4, 6, 5, 4, 4], 
             [5, 4, 4, 5], [10, 4, 4, 4], [4, 4, 5, 4, 4], [4, 9, 4, 4], [4, 6, 4, 5, 5], [5, 4, 4, 4, 6], [4]], [[4, 4, 5, 7, 4], [12, 4, 4], [5, 9, 4, 4], [4, 4]]]
    [ [ [ i for i in  I ] for I in ll ]  for ll in lll ]
    dict(enumerate( 'hello world!')).keys() # dict_keys([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]) 
    li = map(lambda x: x -1 , [3, 6, 9, 12] ) # [2, 5, 8, 11], 全ての要素に処理をする
    (list(li))
