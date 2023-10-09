import openpyxl
import os.path
import itertools
import ls
import sync

class L():
    def kl():
        return ['区分','区間','名称1','号数','始月','始日','始刊','終月','終日','終刊','取扱','メモ2']
    
class a():
    def e(wb,ws):
        ws['B1']=100
        
        ( bool(openpyxl.load_workbook("../chit/内幸町8号朝刊.xlsx")))
        (wb["Sheet1"],wb.worksheets[0],wb.active)
        
        # ワークシートの全セルを反復（行ごと）
        for row in ws.iter_rows():
            for cell in row:
                (cell.value)
        
        # セルの値を取得する
        for row in ws.iter_rows(values_only=True):
            for value in row:
                (value)
                        
        # ワークシートの全セルを反復（列ごと）
        for column in ws.iter_cols(values_only=True):
            (column)
        
        for column in ws.columns:
            for cell in column:
                print(cell.value)
        
    def show( wb, ws):
        # ワークシートの全セルを反復（行ごと）
        for row in ws.iter_rows():
            for cell in row:
                (cell.value)
        
        # セルの値を取得する
        for row in ws.iter_rows(values_only=True):
            for value in row:
                (value)
                        
        # ワークシートの全セルを反復（列ごと）
        for column in ws.iter_cols(values_only=True):
            (column)
        
        for column in ws.columns:
            for cell in column:
                (cell.value)
                
        ws = wb.active  # アクティブなワークシートを選択
        (f'active sheet : {ws.title}')  # sheet name: Sheet
        # ワークシートの列挙
        for sheet in wb:
            (f'sheet : {sheet.title}')
                
    def ph():
        return "../chit/" + sync.dr.area() + ('朝刊' if sync.dr.d_d() != 'dusk' else '夕刊') + ".xlsx"
    
    def rd(lst):
        path = a.ph()
        (sync.dr.area(), '朝刊' if sync.dr.d_d() != 'dusk' else '夕刊' , path)
        wb= openpyxl.load_workbook(path)
        ws = wb["Sheet1"]
        cell1 = ws["A1"].value
        print(ws,cell1)
        #ファイルを閉じる
        wb.close()
        
    def exe():
        pass
        
    def save( book):
        path = a.ph() # ../chit/内幸町8号朝刊.xlsx
        book.save(path) 
        # ファイルを閉じる
        book.close()
        
    def wr( lst):
        path = a.ph()
        wb = openpyxl.load_workbook(path)
        wb.remove(wb.worksheets[-1]) # 'sheet'を削除。
        lt = [ ( c , r ) for r , c in itertools.product( range( 3, 2 * 4, 2), range( 10, 10 + 7 ) )] # [(10, 3), (11, 3), (12, 3), (13, 3), (14, 3), (15, 3), (10, 5), (11, 5), (12, 5), (13, 5), (14, 5), (15, 5), (10, 7), (11, 7), (12, 7), (13, 7), (14, 7), (15, 7), (10, 9), (11, 9), (12, 9), (13, 9), (14, 9), (15, 9)]
        
        for I , dic in enumerate(lst):
            ws = wb.create_sheet( index=-1, title= dic['名称1'])
            a.meta( ws , dic )
            a.name( ws , dic )
            a.temp( ws , dic ) if bool(dic.get('取扱')) else None
            a.memo( ws , dic ) if bool(dic.get('メモ2')) else None
            a.news( lt, ws , dic )
            a.style( ws, wb)
            a.page( wb)
            a.show( wb, ws )
            a.save( wb)
            
    def style( ws, wb):
        # 列の幅を調整(幅)
        for i in ['C','D','E','F','G','H',]:
            ws.column_dimensions[i].width = 44
        # 行の高さを調整
        for i in [  r  for r  in range( 2, 2 + 15 ) ]:
            ws.row_dimensions[i].height = 60
        # フォントサイズを指定
        lt = [ ( c , r ) for r , c in itertools.product( range( 3, 2 * 7), range( 2, 2 + 15 ) )]
        for r,c in lt:
            ws.cell(row=r, column=c).font = openpyxl.styles.Font( size=54) 
        # 文字を縮小して全体を表示する
        lt = [ ( c , r ) for r , c in itertools.product( range( 3, 2 * 7), range( 10, 2 + 15 ) )]
        for r,c in lt:
            pass
            ws.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(shrink_to_fit=True) # 
    
    # 印刷設定
    def page( wb ): 
        print_area = 'C2:K15'
        print_title_rows = '1:5'
        footer_text = '&P / &Nページ'
        for ws in wb.worksheets:
            
            # プリントエリアを設定
            ws.print_area = print_area
            # 常に印刷する行を指定
            ws.print_title_rows = print_title_rows
            # フッターを設定
            ws.oddFooter.center.text = footer_text
        
            wps = ws.page_setup
            # 印刷の向きを設定
            wps.orientation = ws.ORIENTATION_LANDSCAPE
            # 横を１ページ
            wps.fitToWidth = 1
            # 縦を自動
            wps.fitToHeight = 0
            # fitTo属性を有効にする
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            # 用紙サイズを設定
            wps.paperSize = ws.PAPERSIZE_A5 # ws.PAPERSIZE_A5
            # A4サイズ「縦 210mm × 横 297mm」, B5サイズ「182mm × 257mm」
            wps.paperHeight = "257mm"  # 単位：'mm' | 'cm' | 'in' | 'pt' | 'pc' | 'pi'
            wps.paperWidth =  "182mm"  # 単位：'mm' | 'cm' | 'in' | 'pt' | 'pc' | 'pi'
            # 余白
            cm = 1 / 2.54 # 1インチ=2.54cm
            ws.page_margins.top     = 1 * cm
            ws.page_margins.right   = 1 * cm
            ws.page_margins.bottom  = 1 * cm
            ws.page_margins.left    = 1 * cm
            # 印刷品質 ※解像度
            ws.page_setup.horizontalDpi = 1200
            ws.page_setup.verticalDpi = 1200
        
    def meta( ws , dic):
        dct = { k : v for k , v in dic.items() if k in ['区分','区間','号数'] } 
        lt =[(2, 3), (3, 3), (4, 3), (5, 3)]
        for k , v in dct.items():
            r = list(dic.keys()).index( k ) 
            r , c = lt[ r ]
            ws.cell( row = r , column = c ).value = k
            ws.cell( row = r , column = c + 1 ).value = v
            
    def name( ws , dic):
        dct = { k : v for k , v in dic.items() if k in ['名称1'] } 
        for k , v in dct.items():
            r = list(dic.keys()).index( k ) 
            k , v = a.edit( k , v )
            r , c = (4, 3)
            ws.cell( row = r , column = c ).value = k
            ws.cell( row = r , column = c + 1 ).value = v + ' 様'
            
    def edit( k , v ):
        if k in ['名称1']:
            k = '名称'
            v = v 
        return k , v
    
    def temp( ws , dic):
        r, c = ( 6, 3)
        t0, t1, t2, t3, t4, t5, t6 = [ dic.get(i) for i in ls.tm()]
        txt = "【一時止】{0}月{1}日{2}～{3}月{4}日{5}迄　取置：{6}".format( t0, t1, t2, t3, t4, t5, t6)
        ws.cell( row = r , column = c ).value = txt
    
    def memo( ws , dic):
        r, c = ( 8, 3)
        txt, = [ dic.get(i) for i in ['メモ2'] ] # 
        txt = "【メモ】{0}".format(txt)
        ws.cell( row = r , column = c ).value = txt
            
    def news( lt, ws , dic):
        hss ={'ア':'朝日新聞','マ':'毎日新聞','ヨ':'読売新聞','サ':'日本経済新聞','産':'産経新聞','ト':'東京新聞','工':'日刊工業新聞','日産':'日経産業新聞','流通':'日経流通新聞','A':'The Japan Times','Y':'The Japan News','F・T':'Financial Times'}
        d = { k : v for k , v in dic.items() if k not in ls.kl() }   # {'秘書室①': {'ア': '4', 'マ': '5', 'ヨ': '3', 'サ': '3', '産': '4', 'ト': '3'}, '総務課②': {'ア': '1', 'サ': '2'}}
        ( d )
        for k , v in d.items():
            r = list(d.keys()).index( k ) 
            r , c = lt[ r ]
            k = k if k not in hss else hss.get( k )
            k = k[:14]
            v = v + '部'
            ws.cell( row = r , column = c ).value = k
            ws.cell( row = r , column = c + 1 ).value = v 
            
    def active(lst):
        lst = lst[73:78]
        a.rm('../chit/') # フォルダ内の全ファイル削除
        a.create() # ファイル生成
        a.wr( lst)

    def calc(lst):
        lst = lst[33:35]
        dd = { dic.get('名称1'):{ k : v for k , v in dic.items() if k not in ls.kl() } for I , dic in enumerate(lst) } # {'秘書室①': {'ア': '4', 'マ': '5', 'ヨ': '3', 'サ': '3', '産': '4', 'ト': '3'}, '総務課②': {'ア': '1', 'サ': '2'}}
        print(dd)
        
    def rm( ph): # ファイル削除
        l = os.listdir( path = ph )
        [ os.remove( ph + '/' + i ) for i in l ] # ファイル削除
        
    def create(): # "内幸町8号朝刊.xlsx"ファイル生成
        path = a.ph()
        wb = openpyxl.Workbook()
        wb.save(path)
        
if __name__ == '__main__': # これがないと外部からインポートされた際に処理が実行されてしまう。    
    a.calc#(ls.customer())
    a.active( ls.customer())
    a.rd#( ls.customer())
